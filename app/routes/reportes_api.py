"""
Reportes Ejecutivos API - FASE 9
Suite completa de reportes para gerencia

Endpoints:
- GET  /api/v1/reportes/ventas        - Resumen de ventas por período
- GET  /api/v1/reportes/productos     - Top productos por ingresos/margen
- GET  /api/v1/reportes/meseros       - Rendimiento de meseros
- GET  /api/v1/reportes/rentabilidad  - Análisis de rentabilidad
- GET  /api/v1/reportes/inventario    - Análisis de stock (ABC)
- GET  /api/v1/reportes/exportar      - Exportar a Excel
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime, timedelta
from decimal import Decimal
import io
import logging

from app.database import get_db
from sqlalchemy.orm import Session
from sqlalchemy import func, select, and_
from app.models import (
    Venta, DetalleVenta, Producto, Categoria,
    Empleado, Usuario, Empresa
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/v1/reportes",
    tags=["Reportes - FASE 9"]
)


# ============================================================================
# Schemas de respuesta
# ============================================================================

class ReporteVentasResponse(BaseModel):
    periodo_desde: str
    periodo_hasta: str
    total_ventas: float
    cantidad_transacciones: int
    ticket_promedio: float
    total_descuentos: float
    total_impuestos: float
    total_propinas: float
    ventas_por_dia: List[dict]
    ventas_por_medio_pago: List[dict]


class ProductoReporteItem(BaseModel):
    producto_id: int
    codigo: str
    nombre: str
    cantidad_vendida: float
    ingresos: float
    costo_total: float
    margen_bruto: float
    margen_porcentaje: float


class MeseroReporteItem(BaseModel):
    usuario_id: int
    nombre: str
    ventas_cerradas: int
    total_ventas: float
    ticket_promedio: float
    propinas: float


class RentabilidadResponse(BaseModel):
    periodo_desde: str
    periodo_hasta: str
    ingresos_totales: float
    costo_ventas: float
    margen_bruto: float
    margen_bruto_pct: float
    descuentos: float
    propinas: float
    resultado_neto: float


class InventarioReporteItem(BaseModel):
    producto_id: int
    codigo: str
    nombre: str
    existencias: float
    stock_minimo: float
    valor_inventario: float
    estado: str          # ok, bajo, critico, sin_stock
    categoria_abc: str   # A (20% productos, 80% valor), B, C


# ============================================================================
# Helper: parsear fechas con defaults
# ============================================================================

def _rango_fechas(desde: Optional[date], hasta: Optional[date]):
    hoy = date.today()
    d = desde or (hoy - timedelta(days=30))
    h = hasta or hoy
    return d, h


def _empresa_id(db: Session) -> int:
    row = db.query(Empresa).first()
    return row.id if row else 1


# ============================================================================
# GET /ventas
# ============================================================================

@router.get("/ventas", response_model=ReporteVentasResponse)
def reporte_ventas(
    desde: Optional[date] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    hasta: Optional[date] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """Resumen de ventas del período: totales, ticket promedio, medios de pago."""
    d, h = _rango_fechas(desde, hasta)
    empresa_id = _empresa_id(db)

    filtro = and_(
        Venta.empresa_id == empresa_id,
        Venta.estado.in_(["pagada", "credito"]),
        func.date(Venta.fecha) >= d,
        func.date(Venta.fecha) <= h
    )

    ventas = db.query(Venta).filter(filtro).all()

    total_ventas = float(sum(v.total for v in ventas))
    total_descuentos = float(sum(v.descuento or 0 for v in ventas))
    total_impuestos = float(sum(v.impuesto or 0 for v in ventas))
    total_propinas = float(sum(v.propina or 0 for v in ventas))
    cantidad = len(ventas)
    ticket_promedio = total_ventas / cantidad if cantidad else 0.0

    # Ventas por día
    ventas_dia: dict = {}
    for v in ventas:
        dia = v.fecha.date().isoformat()
        ventas_dia[dia] = ventas_dia.get(dia, 0.0) + float(v.total)
    ventas_por_dia = [{"fecha": k, "total": v} for k, v in sorted(ventas_dia.items())]

    # Ventas por medio de pago
    medios: dict = {}
    for v in ventas:
        medio = v.medio_pago or "efectivo"
        medios[medio] = medios.get(medio, 0.0) + float(v.total)
    ventas_por_medio = [{"medio": k, "total": v} for k, v in sorted(medios.items(), key=lambda x: -x[1])]

    logger.info(f"📊 Reporte ventas {d}→{h}: ${total_ventas:,.0f} en {cantidad} transacciones")

    return ReporteVentasResponse(
        periodo_desde=d.isoformat(),
        periodo_hasta=h.isoformat(),
        total_ventas=total_ventas,
        cantidad_transacciones=cantidad,
        ticket_promedio=ticket_promedio,
        total_descuentos=total_descuentos,
        total_impuestos=total_impuestos,
        total_propinas=total_propinas,
        ventas_por_dia=ventas_por_dia,
        ventas_por_medio_pago=ventas_por_medio
    )


# ============================================================================
# GET /productos
# ============================================================================

@router.get("/productos", response_model=List[ProductoReporteItem])
def reporte_productos(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    limite: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """Top productos por ingresos con margen bruto."""
    d, h = _rango_fechas(desde, hasta)
    empresa_id = _empresa_id(db)

    rows = (
        db.query(
            Producto.id,
            Producto.codigo,
            Producto.nombre,
            func.sum(DetalleVenta.cantidad).label("cantidad"),
            func.sum(DetalleVenta.precio * DetalleVenta.cantidad).label("ingresos"),
            func.sum(DetalleVenta.costo_unitario * DetalleVenta.cantidad).label("costo")
        )
        .join(DetalleVenta, DetalleVenta.producto_id == Producto.id)
        .join(Venta, Venta.id == DetalleVenta.venta_id)
        .filter(
            Venta.empresa_id == empresa_id,
            Venta.estado.in_(["pagada", "credito"]),
            func.date(Venta.fecha) >= d,
            func.date(Venta.fecha) <= h
        )
        .group_by(Producto.id, Producto.codigo, Producto.nombre)
        .order_by(func.sum(DetalleVenta.precio * DetalleVenta.cantidad).desc())
        .limit(limite)
        .all()
    )

    resultado = []
    for r in rows:
        ingresos = float(r.ingresos or 0)
        costo = float(r.costo or 0)
        margen = ingresos - costo
        pct = (margen / ingresos * 100) if ingresos else 0.0
        resultado.append(ProductoReporteItem(
            producto_id=r.id,
            codigo=r.codigo,
            nombre=r.nombre,
            cantidad_vendida=float(r.cantidad or 0),
            ingresos=ingresos,
            costo_total=costo,
            margen_bruto=margen,
            margen_porcentaje=round(pct, 2)
        ))

    return resultado


# ============================================================================
# GET /meseros
# ============================================================================

@router.get("/meseros", response_model=List[MeseroReporteItem])
def reporte_meseros(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Rendimiento de meseros: ventas cerradas, total, ticket promedio y propinas."""
    d, h = _rango_fechas(desde, hasta)
    empresa_id = _empresa_id(db)

    rows = (
        db.query(
            Usuario.id,
            Usuario.usuario,
            Empleado.nombre,
            func.count(Venta.id).label("ventas"),
            func.sum(Venta.total).label("total"),
            func.sum(Venta.propina).label("propinas")
        )
        .join(Venta, Venta.usuario_id == Usuario.id)
        .outerjoin(Empleado, Empleado.id == Usuario.empleado_id)
        .filter(
            Venta.empresa_id == empresa_id,
            Venta.estado.in_(["pagada", "credito"]),
            func.date(Venta.fecha) >= d,
            func.date(Venta.fecha) <= h
        )
        .group_by(Usuario.id, Usuario.usuario, Empleado.nombre)
        .order_by(func.sum(Venta.total).desc())
        .all()
    )

    resultado = []
    for r in rows:
        total = float(r.total or 0)
        ventas = int(r.ventas or 0)
        resultado.append(MeseroReporteItem(
            usuario_id=r.id,
            nombre=r.nombre or r.usuario,
            ventas_cerradas=ventas,
            total_ventas=total,
            ticket_promedio=round(total / ventas, 2) if ventas else 0.0,
            propinas=float(r.propinas or 0)
        ))

    return resultado


# ============================================================================
# GET /rentabilidad
# ============================================================================

@router.get("/rentabilidad", response_model=RentabilidadResponse)
def reporte_rentabilidad(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """P&L simplificado: ingresos, costo de ventas, margen bruto."""
    d, h = _rango_fechas(desde, hasta)
    empresa_id = _empresa_id(db)

    filtro = and_(
        Venta.empresa_id == empresa_id,
        Venta.estado.in_(["pagada", "credito"]),
        func.date(Venta.fecha) >= d,
        func.date(Venta.fecha) <= h
    )

    ventas = db.query(Venta).filter(filtro).all()

    ingresos = float(sum(v.total for v in ventas))
    descuentos = float(sum(v.descuento or 0 for v in ventas))
    propinas = float(sum(v.propina or 0 for v in ventas))

    # Costo real desde detalles
    costo_ventas = float(
        db.query(func.coalesce(func.sum(DetalleVenta.costo_unitario * DetalleVenta.cantidad), 0))
        .join(Venta, Venta.id == DetalleVenta.venta_id)
        .filter(filtro)
        .scalar() or 0
    )

    margen_bruto = ingresos - costo_ventas
    margen_pct = round((margen_bruto / ingresos * 100) if ingresos else 0.0, 2)
    resultado_neto = margen_bruto - descuentos + propinas

    return RentabilidadResponse(
        periodo_desde=d.isoformat(),
        periodo_hasta=h.isoformat(),
        ingresos_totales=ingresos,
        costo_ventas=costo_ventas,
        margen_bruto=margen_bruto,
        margen_bruto_pct=margen_pct,
        descuentos=descuentos,
        propinas=propinas,
        resultado_neto=resultado_neto
    )


# ============================================================================
# GET /inventario
# ============================================================================

@router.get("/inventario", response_model=List[InventarioReporteItem])
def reporte_inventario(
    solo_bajos: bool = Query(False, description="Mostrar solo productos con stock bajo"),
    db: Session = Depends(get_db)
):
    """Análisis de inventario con clasificación ABC por valor."""
    empresa_id = _empresa_id(db)

    query = db.query(Producto).filter(
        Producto.empresa_id == empresa_id,
        Producto.activo == True
    )
    productos = query.all()

    if not productos:
        return []

    # Calcular valor de inventario de cada producto
    items_con_valor = []
    for p in productos:
        valor = float(p.existencias * p.costo)
        items_con_valor.append((p, valor))

    # Clasificación ABC por valor de inventario
    total_valor = sum(v for _, v in items_con_valor)
    items_con_valor.sort(key=lambda x: -x[1])

    acumulado = 0.0
    resultado = []
    for p, valor in items_con_valor:
        if total_valor > 0:
            acumulado += valor
            pct_acum = acumulado / total_valor
            if pct_acum <= 0.80:
                cat_abc = "A"
            elif pct_acum <= 0.95:
                cat_abc = "B"
            else:
                cat_abc = "C"
        else:
            cat_abc = "C"

        # Estado de stock
        existencias = float(p.existencias)
        minimo = float(p.stock_minimo)
        if existencias <= 0:
            estado = "sin_stock"
        elif minimo > 0 and existencias <= minimo * 0.5:
            estado = "critico"
        elif minimo > 0 and existencias <= minimo:
            estado = "bajo"
        else:
            estado = "ok"

        if solo_bajos and estado == "ok":
            continue

        resultado.append(InventarioReporteItem(
            producto_id=p.id,
            codigo=p.codigo,
            nombre=p.nombre,
            existencias=existencias,
            stock_minimo=minimo,
            valor_inventario=valor,
            estado=estado,
            categoria_abc=cat_abc
        ))

    return resultado


# ============================================================================
# GET /exportar  — Excel
# ============================================================================

@router.get("/exportar")
def exportar_reporte(
    tipo: str = Query(..., description="ventas | productos | inventario | meseros"),
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Exporta cualquier reporte a Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado. Ejecute: pip install openpyxl")

    d, h = _rango_fechas(desde, hasta)
    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="B45309")
    header_align = Alignment(horizontal="center")

    def _set_headers(cols):
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    tipos_validos = {"ventas", "productos", "inventario", "meseros"}
    if tipo not in tipos_validos:
        raise HTTPException(400, f"Tipo inválido. Use: {', '.join(tipos_validos)}")

    if tipo == "ventas":
        ws.title = "Ventas"
        datos = reporte_ventas(desde=d, hasta=h, db=db)
        _set_headers(["Fecha", "Total ($)"])
        for row in datos.ventas_por_dia:
            ws.append([row["fecha"], row["total"]])

    elif tipo == "productos":
        ws.title = "Productos"
        datos = reporte_productos(desde=d, hasta=h, limite=100, db=db)
        _set_headers(["Código", "Nombre", "Cant. Vendida", "Ingresos", "Costo", "Margen", "Margen %"])
        for p in datos:
            ws.append([p.codigo, p.nombre, p.cantidad_vendida, p.ingresos, p.costo_total, p.margen_bruto, p.margen_porcentaje])

    elif tipo == "inventario":
        ws.title = "Inventario"
        datos = reporte_inventario(solo_bajos=False, db=db)
        _set_headers(["Código", "Nombre", "Existencias", "Mínimo", "Valor ($)", "Estado", "ABC"])
        for p in datos:
            ws.append([p.codigo, p.nombre, p.existencias, p.stock_minimo, p.valor_inventario, p.estado, p.categoria_abc])

    elif tipo == "meseros":
        ws.title = "Meseros"
        datos = reporte_meseros(desde=d, hasta=h, db=db)
        _set_headers(["Nombre", "Ventas", "Total ($)", "Ticket Promedio", "Propinas"])
        for m in datos:
            ws.append([m.nombre, m.ventas_cerradas, m.total_ventas, m.ticket_promedio, m.propinas])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reporte_{tipo}_{d}_{h}.xlsx"
    logger.info(f"📥 Exportando reporte {tipo} {d}→{h}")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
